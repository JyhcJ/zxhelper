import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 读取 Excel 文件
input_file = "input.xlsx"  # 输入文件名
output_file = "output.xlsx"  # 输出文件名

# 读取数据
df = pd.read_excel(input_file, header=None)  # 假设没有表头

# 找到需要重新排序的起始行
start_index = df[df[0] == 1].index[0]  # 找到序号为 1 的行
start_number = 1  # 重新排序的起始序号

# 重新排序
for i in range(start_index, len(df)):
    if pd.notna(df.iloc[i, 0]):  # 如果当前行有序号
        df.iloc[i, 0] = start_number
        start_number += 1

# 保存到新的 Excel 文件
df.to_excel(output_file, index=False, header=False)

# 使用 openpyxl 合并空白单元格
wb = load_workbook(output_file)
ws = wb.active

# 动态合并空白单元格
i = 2  # 从第二行开始（openpyxl 行号从 1 开始）
while i <= ws.max_row:
    if ws.cell(row=i, column=1).value is None:  # 如果当前单元格为空
        merge_start = i - 1
        while i <= ws.max_row and ws.cell(row=i, column=1).value is None:
            i += 1
        merge_end = i - 1
        
        # 合并单元格
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)
        
        # 设置合并单元格的对齐方式
        merged_cell = ws.cell(row=merge_start, column=1)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        i += 1

# 保存修改后的文件
wb.save(output_file)

print(f"处理完成，结果已保存到 {output_file}")